"""Export functionality for tickets, questions, and test cases."""

import json
from datetime import datetime
from io import BytesIO
from typing import Optional

from app.db.model import Ticket, GeneratedContent, SessionLocal


def export_tickets_to_excel(
    tickets: list[Ticket],
    include_questions: bool = True,
    include_tests: bool = True
) -> BytesIO:
    """
    Export tickets to Excel format.

    Returns:
        BytesIO buffer containing the Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Tickets sheet
    ws_tickets = wb.active
    ws_tickets.title = "Tickets"

    ticket_headers = ["Key", "Title", "Type", "Status", "Quality Score", "Updated", "Questions", "Test Cases"]
    for col, header in enumerate(ticket_headers, 1):
        cell = ws_tickets.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    session = SessionLocal()
    try:
        for row, ticket in enumerate(tickets, 2):
            # Get generated content counts
            questions_count = session.query(GeneratedContent).filter_by(
                ticket_id=ticket.id, content_type='questions'
            ).count()
            tests_count = session.query(GeneratedContent).filter_by(
                ticket_id=ticket.id, content_type='test_cases'
            ).count()

            values = [
                ticket.jira_key,
                ticket.title or "",
                ticket.issue_type or "",
                ticket.status or "",
                ticket.quality_score if ticket.quality_score else "",
                ticket.updated_at.strftime("%Y-%m-%d") if ticket.updated_at else "",
                "Yes" if questions_count > 0 else "No",
                "Yes" if tests_count > 0 else "No"
            ]

            for col, value in enumerate(values, 1):
                cell = ws_tickets.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Adjust column widths
        ws_tickets.column_dimensions['A'].width = 15
        ws_tickets.column_dimensions['B'].width = 50
        ws_tickets.column_dimensions['C'].width = 12
        ws_tickets.column_dimensions['D'].width = 15
        ws_tickets.column_dimensions['E'].width = 12
        ws_tickets.column_dimensions['F'].width = 12
        ws_tickets.column_dimensions['G'].width = 12
        ws_tickets.column_dimensions['H'].width = 12

        # Questions sheet
        if include_questions:
            ws_questions = wb.create_sheet("Questions")
            q_headers = ["Ticket Key", "Ticket Title", "Questions", "Published", "Generated At"]
            for col, header in enumerate(q_headers, 1):
                cell = ws_questions.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            row = 2
            for ticket in tickets:
                contents = session.query(GeneratedContent).filter_by(
                    ticket_id=ticket.id, content_type='questions'
                ).all()

                for content in contents:
                    try:
                        questions_data = json.loads(content.content)
                        questions_text = "\n".join(f"• {q}" for q in questions_data.get('questions', []))
                    except:
                        questions_text = content.content

                    values = [
                        ticket.jira_key,
                        ticket.title or "",
                        questions_text,
                        "Yes" if content.published else "No",
                        content.created_at.strftime("%Y-%m-%d %H:%M") if content.created_at else ""
                    ]

                    for col, value in enumerate(values, 1):
                        cell = ws_questions.cell(row=row, column=col, value=value)
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                    row += 1

            ws_questions.column_dimensions['A'].width = 15
            ws_questions.column_dimensions['B'].width = 40
            ws_questions.column_dimensions['C'].width = 80
            ws_questions.column_dimensions['D'].width = 12
            ws_questions.column_dimensions['E'].width = 18

        # Test Cases sheet
        if include_tests:
            ws_tests = wb.create_sheet("Test Cases")
            t_headers = ["Ticket Key", "Ticket Title", "Test Cases", "Published", "Generated At"]
            for col, header in enumerate(t_headers, 1):
                cell = ws_tests.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            row = 2
            for ticket in tickets:
                contents = session.query(GeneratedContent).filter_by(
                    ticket_id=ticket.id, content_type='test_cases'
                ).all()

                for content in contents:
                    try:
                        tests_data = json.loads(content.content)
                        tests_text = "\n\n".join(tests_data.get('test_cases', []))
                    except:
                        tests_text = content.content

                    values = [
                        ticket.jira_key,
                        ticket.title or "",
                        tests_text,
                        "Yes" if content.published else "No",
                        content.created_at.strftime("%Y-%m-%d %H:%M") if content.created_at else ""
                    ]

                    for col, value in enumerate(values, 1):
                        cell = ws_tests.cell(row=row, column=col, value=value)
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                    row += 1

            ws_tests.column_dimensions['A'].width = 15
            ws_tests.column_dimensions['B'].width = 40
            ws_tests.column_dimensions['C'].width = 100
            ws_tests.column_dimensions['D'].width = 12
            ws_tests.column_dimensions['E'].width = 18

    finally:
        session.close()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_ticket_to_markdown(ticket: Ticket) -> str:
    """Export a single ticket with all generated content to Markdown."""
    session = SessionLocal()
    try:
        lines = [
            f"# {ticket.jira_key}: {ticket.title or 'No Title'}",
            "",
            f"**Type:** {ticket.issue_type or 'Unknown'}",
            f"**Status:** {ticket.status or 'Unknown'}",
        ]

        if ticket.quality_score:
            lines.append(f"**Quality Score:** {ticket.quality_score}/10")

        if ticket.updated_at:
            lines.append(f"**Last Updated:** {ticket.updated_at.strftime('%Y-%m-%d %H:%M')}")

        lines.extend(["", "## Description", ""])
        lines.append(ticket.description or "_No description provided_")

        # Quality analysis
        if ticket.quality_summary:
            lines.extend(["", "## Quality Analysis", ""])
            lines.append(f"**Summary:** {ticket.quality_summary}")

            if ticket.quality_issues:
                try:
                    issues = json.loads(ticket.quality_issues)
                    if issues:
                        lines.extend(["", "**Issues:**"])
                        for issue in issues:
                            lines.append(f"- {issue}")
                except:
                    pass

            if ticket.quality_suggestions:
                try:
                    suggestions = json.loads(ticket.quality_suggestions)
                    if suggestions:
                        lines.extend(["", "**Suggestions:**"])
                        for suggestion in suggestions:
                            lines.append(f"- {suggestion}")
                except:
                    pass

        # Questions
        questions_content = session.query(GeneratedContent).filter_by(
            ticket_id=ticket.id, content_type='questions'
        ).order_by(GeneratedContent.created_at.desc()).first()

        if questions_content:
            lines.extend(["", "## Refinement Questions", ""])
            try:
                data = json.loads(questions_content.content)
                for q in data.get('questions', []):
                    lines.append(f"- {q}")
            except:
                lines.append(questions_content.content)

            status = "Published" if questions_content.published else "Not Published"
            lines.append(f"\n_Status: {status}_")

        # Test Cases
        tests_content = session.query(GeneratedContent).filter_by(
            ticket_id=ticket.id, content_type='test_cases'
        ).order_by(GeneratedContent.created_at.desc()).first()

        if tests_content:
            lines.extend(["", "## Test Cases", ""])
            try:
                data = json.loads(tests_content.content)
                for tc in data.get('test_cases', []):
                    lines.append(tc)
                    lines.append("")
            except:
                lines.append(tests_content.content)

            status = "Published" if tests_content.published else "Not Published"
            lines.append(f"_Status: {status}_")

        lines.extend(["", "---", f"_Generated by ProRef on {datetime.now().strftime('%Y-%m-%d %H:%M')}_"])

        return "\n".join(lines)

    finally:
        session.close()


def export_sprint_report_markdown(tickets: list[Ticket]) -> str:
    """Generate a sprint report in Markdown format."""
    session = SessionLocal()
    try:
        # Statistics
        total = len(tickets)
        with_questions = sum(1 for t in tickets if t.questions_generated)
        with_tests = sum(1 for t in tickets if t.test_cases_generated)

        questions_published = 0
        tests_published = 0

        for ticket in tickets:
            q = session.query(GeneratedContent).filter_by(
                ticket_id=ticket.id, content_type='questions', published=True
            ).first()
            if q:
                questions_published += 1

            t = session.query(GeneratedContent).filter_by(
                ticket_id=ticket.id, content_type='test_cases', published=True
            ).first()
            if t:
                tests_published += 1

        # Quality distribution
        quality_ready = sum(1 for t in tickets if t.quality_score and t.quality_score >= 8)
        quality_needs_work = sum(1 for t in tickets if t.quality_score and 5 <= t.quality_score < 8)
        quality_not_ready = sum(1 for t in tickets if t.quality_score and t.quality_score < 5)
        quality_unscored = sum(1 for t in tickets if not t.quality_score)

        # Build report
        lines = [
            "# Sprint Refinement Report",
            f"_Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}_",
            "",
            "## Overview",
            "",
            f"| Metric | Count |",
            f"|--------|-------|",
            f"| Total Tickets | {total} |",
            f"| With Questions | {with_questions} |",
            f"| With Test Cases | {with_tests} |",
            f"| Questions Published | {questions_published} |",
            f"| Test Cases Published | {tests_published} |",
            "",
            "## Quality Distribution",
            "",
            f"| Status | Count |",
            f"|--------|-------|",
            f"| Ready (8-10) | {quality_ready} |",
            f"| Needs Work (5-7) | {quality_needs_work} |",
            f"| Not Ready (1-4) | {quality_not_ready} |",
            f"| Not Scored | {quality_unscored} |",
            "",
            "## Tickets by Status",
            ""
        ]

        # Group by status
        status_groups = {}
        for ticket in tickets:
            status = ticket.status or "Unknown"
            if status not in status_groups:
                status_groups[status] = []
            status_groups[status].append(ticket)

        for status, group in sorted(status_groups.items()):
            lines.append(f"### {status} ({len(group)})")
            lines.append("")
            for ticket in group:
                score_badge = f" [Score: {ticket.quality_score}]" if ticket.quality_score else ""
                flags = []
                if ticket.questions_generated:
                    flags.append("Q")
                if ticket.test_cases_generated:
                    flags.append("T")
                if ticket.content_changed:
                    flags.append("CHANGED")
                flags_str = f" ({', '.join(flags)})" if flags else ""
                lines.append(f"- **{ticket.jira_key}**: {ticket.title or 'No title'}{score_badge}{flags_str}")
            lines.append("")

        # Tickets needing attention
        needs_attention = [t for t in tickets if t.quality_score and t.quality_score < 5]
        if needs_attention:
            lines.extend([
                "## Tickets Needing Attention",
                "",
                "_These tickets have quality scores below 5 and may need refinement:_",
                ""
            ])
            for ticket in needs_attention:
                lines.append(f"### {ticket.jira_key}: {ticket.title}")
                if ticket.quality_summary:
                    lines.append(f"**Issue:** {ticket.quality_summary}")
                if ticket.quality_suggestions:
                    try:
                        suggestions = json.loads(ticket.quality_suggestions)
                        if suggestions:
                            lines.append("**Suggestions:**")
                            for s in suggestions:
                                lines.append(f"- {s}")
                    except:
                        pass
                lines.append("")

        # Changed tickets
        changed_tickets = [t for t in tickets if t.content_changed]
        if changed_tickets:
            lines.extend([
                "## Changed Tickets",
                "",
                "_These tickets have been modified since content was generated:_",
                ""
            ])
            for ticket in changed_tickets:
                lines.append(f"- **{ticket.jira_key}**: {ticket.title or 'No title'}")
            lines.append("")

        lines.extend([
            "---",
            "_Report generated by ProRef_"
        ])

        return "\n".join(lines)

    finally:
        session.close()
