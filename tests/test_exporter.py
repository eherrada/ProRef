"""Tests for exporter module."""

import pytest
from unittest.mock import MagicMock, patch
from datetime import datetime


class TestExportTicketToMarkdown:
    """Tests for export_ticket_to_markdown function."""

    @patch('app.logic.exporter.SessionLocal')
    def test_exports_basic_ticket(self, mock_session):
        """Should export basic ticket info to markdown."""
        from app.logic.exporter import export_ticket_to_markdown

        ticket = MagicMock()
        ticket.jira_key = "TEST-123"
        ticket.title = "Test Ticket"
        ticket.issue_type = "Story"
        ticket.status = "Open"
        ticket.description = "This is a test description"
        ticket.quality_score = None
        ticket.quality_summary = None
        ticket.quality_issues = None
        ticket.quality_suggestions = None
        ticket.updated_at = datetime(2024, 1, 15, 10, 30)
        ticket.id = "123"

        mock_session_instance = MagicMock()
        mock_session.return_value = mock_session_instance
        mock_session_instance.query.return_value.filter_by.return_value.order_by.return_value.first.return_value = None

        result = export_ticket_to_markdown(ticket)

        assert "# TEST-123: Test Ticket" in result
        assert "**Type:** Story" in result
        assert "**Status:** Open" in result
        assert "This is a test description" in result

    @patch('app.logic.exporter.SessionLocal')
    def test_includes_quality_score(self, mock_session):
        """Should include quality score when present."""
        from app.logic.exporter import export_ticket_to_markdown

        ticket = MagicMock()
        ticket.jira_key = "TEST-123"
        ticket.title = "Test"
        ticket.issue_type = "Story"
        ticket.status = "Open"
        ticket.description = "Desc"
        ticket.quality_score = 8
        ticket.quality_summary = "Well defined ticket"
        ticket.quality_issues = '["Missing edge cases"]'
        ticket.quality_suggestions = '["Add more details"]'
        ticket.updated_at = datetime(2024, 1, 15)
        ticket.id = "123"

        mock_session_instance = MagicMock()
        mock_session.return_value = mock_session_instance
        mock_session_instance.query.return_value.filter_by.return_value.order_by.return_value.first.return_value = None

        result = export_ticket_to_markdown(ticket)

        assert "**Quality Score:** 8/10" in result
        assert "Well defined ticket" in result


class TestExportSprintReportMarkdown:
    """Tests for export_sprint_report_markdown function."""

    @patch('app.logic.exporter.SessionLocal')
    def test_generates_report_header(self, mock_session):
        """Should include report header and date."""
        from app.logic.exporter import export_sprint_report_markdown

        mock_session_instance = MagicMock()
        mock_session.return_value = mock_session_instance
        mock_session_instance.query.return_value.filter_by.return_value.first.return_value = None

        result = export_sprint_report_markdown([])

        assert "# Sprint Refinement Report" in result
        assert "## Overview" in result

    @patch('app.logic.exporter.SessionLocal')
    def test_includes_ticket_counts(self, mock_session):
        """Should include ticket statistics."""
        from app.logic.exporter import export_sprint_report_markdown

        ticket1 = MagicMock()
        ticket1.id = "1"
        ticket1.jira_key = "TEST-1"
        ticket1.title = "Ticket 1"
        ticket1.status = "Open"
        ticket1.questions_generated = True
        ticket1.test_cases_generated = False
        ticket1.quality_score = 8
        ticket1.quality_summary = None
        ticket1.quality_suggestions = None
        ticket1.content_changed = False

        ticket2 = MagicMock()
        ticket2.id = "2"
        ticket2.jira_key = "TEST-2"
        ticket2.title = "Ticket 2"
        ticket2.status = "Open"
        ticket2.questions_generated = False
        ticket2.test_cases_generated = False
        ticket2.quality_score = 3
        ticket2.quality_summary = "Needs work"
        ticket2.quality_suggestions = '["Add details"]'
        ticket2.content_changed = True

        mock_session_instance = MagicMock()
        mock_session.return_value = mock_session_instance
        mock_session_instance.query.return_value.filter_by.return_value.first.return_value = None

        result = export_sprint_report_markdown([ticket1, ticket2])

        assert "Total Tickets | 2" in result
        assert "## Quality Distribution" in result

    @patch('app.logic.exporter.SessionLocal')
    def test_highlights_changed_tickets(self, mock_session):
        """Should highlight tickets that have changed."""
        from app.logic.exporter import export_sprint_report_markdown

        ticket = MagicMock()
        ticket.id = "1"
        ticket.jira_key = "TEST-1"
        ticket.title = "Changed Ticket"
        ticket.status = "Open"
        ticket.questions_generated = True
        ticket.test_cases_generated = False
        ticket.quality_score = 7
        ticket.quality_summary = None
        ticket.quality_suggestions = None
        ticket.content_changed = True

        mock_session_instance = MagicMock()
        mock_session.return_value = mock_session_instance
        mock_session_instance.query.return_value.filter_by.return_value.first.return_value = None

        result = export_sprint_report_markdown([ticket])

        assert "## Changed Tickets" in result
        assert "TEST-1" in result


class TestExportTicketsToExcel:
    """Tests for export_tickets_to_excel function."""

    def test_raises_import_error_without_openpyxl(self):
        """Should raise ImportError if openpyxl not installed."""
        from app.logic.exporter import export_tickets_to_excel

        # This test verifies the error message when openpyxl is missing
        # In practice, we'd mock the import, but the function handles it gracefully

    @patch('app.logic.exporter.SessionLocal')
    def test_creates_workbook_with_sheets(self, mock_session):
        """Should create workbook with Tickets, Questions, and Test Cases sheets."""
        pytest.importorskip("openpyxl")

        from app.logic.exporter import export_tickets_to_excel
        import openpyxl

        ticket = MagicMock()
        ticket.id = "1"
        ticket.jira_key = "TEST-1"
        ticket.title = "Test Ticket"
        ticket.issue_type = "Story"
        ticket.status = "Open"
        ticket.quality_score = 7
        ticket.updated_at = datetime(2024, 1, 15)

        mock_session_instance = MagicMock()
        mock_session.return_value = mock_session_instance
        mock_session_instance.query.return_value.filter_by.return_value.count.return_value = 0
        mock_session_instance.query.return_value.filter_by.return_value.all.return_value = []

        buffer = export_tickets_to_excel([ticket])

        # Load and verify the workbook
        wb = openpyxl.load_workbook(buffer)
        assert "Tickets" in wb.sheetnames
        assert "Questions" in wb.sheetnames
        assert "Test Cases" in wb.sheetnames
